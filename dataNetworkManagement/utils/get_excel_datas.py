#!/usr/bin/python3
# -*- coding:utf-8 -*-
import xlrd
import pandas as pd
import openpyxl
from openpyxl import load_workbook


#  1：Python内置库xlrd，仅支持xls，不支持xlsx
# # 打开文件，获取文件流
# workbook = xlrd.open_workbook('/Users/liyong/Downloads/1、基本粒子/1、标品/4、DDE/测试工作安排计划表.xlsx')
# # 获取文件对应sheet
# sheet = workbook.sheet_by_index(2)
# # 遍历Excel文件指定sheet的每一行，并获取每一行的value赋值给row
# for row_index in range(sheet.nrows):
#     row = sheet.row_values(row_index)
#     print(row)

#  2：Python内置库pandas
# # 打开文件
# df = pd.read_excel('/Users/liyong/Downloads/1、基本粒子/1、标品/4、DDE/测试工作安排计划表.xlsx')
# # 获取默认第一个sheet的前五行数据
# print(df.head())

#  3：Python内置库openpyxl
#  3-a：默认获取第一个sheet
# workbook = load_workbook('/Users/liyong/Downloads/1、基本粒子/1、标品/4、DDE/测试工作安排计划表.xlsx')
# sheet = workbook.active
# for row in sheet.iter_rows():
#     row_data = [cell.value for cell in row]
#     print(row_data)
#  3-b：


class GetExcelDatas():
    def get_excel_datas(self):
        # 打开文件
        excel_file = openpyxl.load_workbook("/Users/liyong/Downloads/1、基本粒子/1、标品/4、DDE/测试工作安排计划表.xlsx")
        # 根据sheet名称获取指定sheet
        excel_sheet = excel_file['数据网络管理系统V0.3自动化测试用例']
        # 获取指定单元格的值方法一
        # cell_value = excel_sheet.cell(row=3, column=6).value
        # 获取指定单元格的值方法二
        # 定义空列表，将下方遍历得到的数据行（以列表形式返回每一行数据），依次添加到该列表，最终得到一个列表嵌套列表
        excel_datas = []
        # 遍历当前sheet的每一行
        for row in excel_sheet.iter_rows():
            # 使用列表推导式获取每一行每个单元格的value，返回一个列表
            row_data = [cell.value for cell in row]
            # 将所有行的数据插入定义好的列表excel_datas中，最终得到列表嵌套列表，每个子列表为一行数据
            excel_datas.append(row_data)
        # 获取第三行第六列的单元格的数据，以string格式返回
        # print(excel_datas[2][5])
        return excel_datas
